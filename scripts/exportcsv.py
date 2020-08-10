#!/usr/bin/env python3

# TODO: Fix quote inconsistency.
# TODO: "graph"
# TODO: CI

# Requires Python >= 3.8 for assignment within condition: https://www.python.org/dev/peps/pep-0572/.
# Also requires Python >= 3.6 for preserving order of inseration for dictionaries.
# Also requires Python >= 3.5 for recursive support in glob.
# pip install -r requirements.txt

# Compatibility (TODO: update this!):
# - The online version of Microsoft Excel works without any issues.
# - Google Sheets has issues with decimal places. The default level of precision is not enough to
# support the math done. This may be fixable.
# - LibreOffice Calc has issues with sheet references. It uses a different format than Excel and
# Google Sheets, but seems to convert the formula to all lowercase, breaking the name of the sheet
# being reference. It is not supported by this script.

import argparse
import csv
from glob import glob
import os
import re

from colorama import Fore, Style, init as colorama_init
import openpyxl

# TODO: Describe these.
DEFAULT_DATA_DIR = "../data"
VOLTAGE_FILENAME_STR = "voltage"
LIFETIME_FILENAME_STR = "lifetime"
DENSITY_FILENAME_STR = "ehpdensity"
# The number of rows for a table in the all integrals sheet which are not used for data. This is
# used to offset the relative pool row numbers.
NUM_NON_DATA_ROWS = 2
# Number of blank rows/columns to insert between tables and charts.
NUM_PADDING_CELLS = 1
NUM_NON_DATA_COLUMNS = 2
ALLINT_VOLTAGE_COL = 1
ALLINT_LIFETIME_COL = 2
DIODE_DENSITY_COL = 1


def sort_alphanumeric(data):
  """See: https://stackoverflow.com/a/48030307
  """

  def convert(text): return int(text) if text.isdigit() else text.lower()
  def alphanum_key(key): return [convert(c)
                                 for c in re.split('([0-9]+)', key)]
  return sorted(data, key=alphanum_key)


def ls(args):
  """TODO
  """

  print(Fore.GREEN + "Listing experiments...")
  # Fetch top-level directories in the data directory.
  child_dirs = glob(args.dir + "/*/")
  # Only consider directories with CSV files in them.
  experiment_dirs = sort_alphanumeric(
      [d for d in child_dirs if glob(d + "/**/*.csv", recursive=True)])
  if experiment_dirs:
    print(Fore.CYAN + "Experiments:")
    for d in experiment_dirs:
      # The path must be normalized in order for basename to not return an empty string.
      print("- " + Fore.WHITE + os.path.basename(os.path.normpath(d)))
    print(Fore.GREEN + "Success.")
    os.sys.exit(0)
  else:
    print(Fore.RED + "No experiments found.")
    os.sys.exit(1)


def export(args):
  """TODO
  """

  print(Fore.GREEN + "Mapping CSV files to dictionary...")

  exp = args.exp
  # Provide support for only passing the name of the experiment.
  if not os.path.exists(exp):
    exp = os.path.join(DEFAULT_DATA_DIR, exp)
    if not os.path.exists(exp):
      print(Fore.RED + "Experiment \"{}\" not found.".format(args.exp))
      os.sys.exit(1)
  csv_list = glob(exp + "/**/*.csv", recursive=True)
  if not csv_list:
    print(Fore.RED + "No CSV files found.")
    os.sys.exit(1)

  print(Fore.CYAN + "CSV Files:")

  # Force exponential notation with 1 sigfig, strip leading 0s in the exponent of exponential
  # notation, and strip unneeded "+".
  def shorten_exp_notation(s):
    return re.sub(r'e(\+)?(-)?0+(.+)', r'e\2\3', format(s, "1.0e"))

  csv_dict = {}
  for f in csv_list:
    print(Fore.WHITE + Style.BRIGHT + f)
    # File names are expected to come in formats like
    # "PC_lifetime0.001_ehpdensity0.0001_time.csv". Here, we parse the information given,
    # delimited by the underscores.
    basename_parts = re.split(r'_|\.csv', os.path.basename(f))

    # See: https://stackoverflow.com/a/30197797.
    device_str = next((s for s in basename_parts if "PC" in s), None)
    if not device_str:
      device_str = next(
          (s for s in basename_parts if "Diode" in s), None)
      if not device_str:
        print(
            Fore.RED +
            "Device name could not be extracted from file \"{}\". Looked for \"PC\" or \"Diode\" \
in file name. Skipping this file.".format(f))
        continue

    # This works in the same way as its voltage analog.
    voltage_str = next(
        (s for s in basename_parts if VOLTAGE_FILENAME_STR in s), None)
    if not voltage_str:
      print(
          Fore.RED +
          "Density could not be extracted from file \"{}\". Looked for \"{}\" in file name. \
Skipping this file." .format(f, VOLTAGE_FILENAME_STR))
      continue
    # TODO: what formatting to use?
    voltage_key = "{}V".format(
        int(voltage_str.replace(VOLTAGE_FILENAME_STR, '')))

    # Look for LIFETIME_FILENAME_STR in the file name to find the field, and then remove that
    # identifier string so that we are just left with the value of the field. For example,
    # "lifetime0.001" is transformed into "τ=1.00e-3".
    lifetime_str = next(
        (s for s in basename_parts if LIFETIME_FILENAME_STR in s), None)
    if not lifetime_str:
      print(
          Fore.RED +
          "Lifetime could not be extracted from file \"{}\". Looked for \"{}\" in file name. \
Skipping this file." .format(f, LIFETIME_FILENAME_STR))
      continue
    # Force scientific notation so that the dictionary key can be properly sorted as a float.
    lifetime_key = "τ={}".format(
        shorten_exp_notation(float(lifetime_str.replace(LIFETIME_FILENAME_STR, ''))))

    # This works in the same way as its lifetime analog.
    density_str = next(
        (s for s in basename_parts if DENSITY_FILENAME_STR in s), None)
    if not density_str:
      print(
          Fore.RED +
          "Density could not be extracted from file \"{}\". Looked for \"{}\" in file name. \
Skipping this file." .format(f, DENSITY_FILENAME_STR))
      continue
    density_key = "D={}".format(
        shorten_exp_notation(float(density_str.replace(DENSITY_FILENAME_STR, ''))))

    # device_str is already usable as a key.
    if device_str not in csv_dict:
      csv_dict[device_str] = {}
    # Initialize lifetime and voltage dictionaries as needed.
    if voltage_key not in csv_dict[device_str]:
      csv_dict[device_str][voltage_key] = {}
    if lifetime_key not in csv_dict[device_str][voltage_key]:
      csv_dict[device_str][voltage_key][lifetime_key] = {}
    csv_dict[device_str][voltage_key][lifetime_key][density_key] = f

  if not csv_dict:
    print(Fore.RED + "Unable to create a CSV dictionary.")
    os.sys.exit(1)
  print(Fore.GREEN + "Success.")

  # This is necessary because alphanumeric sorting of the file names may not account for
  # scientific notation.
  print(Fore.GREEN + "Ordering dictionary...")

  # This (https://stackoverflow.com/a/47882384) SO answer explains how to recursively sort a
  # dictionary. I have modified this to sort the root entries as normal, but sort further nested
  # entries as floats/ints. This also accounts for the equal sign in the key. This was annoying to
  # figure out.
  def sort_dict(dictionary, level=0):
    # print("sort_dict (level {}), {}".format(level, dictionary))
    result = {}
    # While sorting, ignore .

    def key(x):
      if "V" in x[0]:
        return int(x[0].split('V')[0])
      elif "=" in x[0]:
        return float(x[0].split('=')[-1])
      else:
        return x
    for k, v in sorted(dictionary.items(), key=key):
      # print("{}:{}".format(k, v))
      if isinstance(v, dict):
        result[k] = sort_dict(v, level + 1)
      else:
        result[k] = v
    return result

  csv_dict_ordered = sort_dict(csv_dict)

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Creating row and column mappings.")

  # Initialize the pool. TODO: Document this!
  # These are relative!
  pool = {"row_mappings": {}, "col_mappings": {}}

  # Initiialize the pool keys.
  for device in csv_dict_ordered:
    for voltage in csv_dict_ordered[device]:
      pool["row_mappings"][voltage] = {}
      for lifetime in csv_dict_ordered[device][voltage]:
        pool["row_mappings"][voltage][lifetime] = {}
        for density in csv_dict_ordered[device][voltage][lifetime]:
          pool["col_mappings"][density] = {}

  # Sort the pool by its keys.
  pool = sort_dict(pool)

  # Initialize the pool values.
  current_row = 1
  for voltage in pool["row_mappings"]:
    for lifetime in pool["row_mappings"][voltage]:
      pool["row_mappings"][voltage][lifetime] = current_row
      current_row += 1
  current_col = 1
  for density in pool["col_mappings"]:
    pool["col_mappings"][density] = current_col
    current_col += 1

  # This function takes a dictionary, goes through each nested dictionary, and finds the value that
  # best satisfies function f.
  # See: https://stackoverflow.com/a/280156.
  def find_dict_val(dic, f):
    first_sub_dic = dic[next(iter(dic))]
    # Initialize to first value in dict.
    val = first_sub_dic[f(first_sub_dic, key=first_sub_dic.get)]
    for sub_dic_key in dic:
      sub_dic = dic[sub_dic_key]
      val = f(val, sub_dic[f(sub_dic, key=sub_dic.get)])
    return val

  # Initialize the device pool, providing the .
  # These are absolute!
  device_pool = {}
  device_n = 0
  # These two variables are relative to the first row (index 1 - spreadsheets aren't zero-indexed!)
  # of the data portion of the table.
  #
  # This one should just be 1.
  relative_row_data_start = find_dict_val(pool["row_mappings"], min)
  relative_row_data_end = find_dict_val(pool["row_mappings"], max)
  num_data_rows = relative_row_data_end - relative_row_data_start
  # Add another device so that we can have a table for normalized Photoconductor values.
  csv_dict_ordered["PC_Norm"] = {}
  for device in csv_dict_ordered:
    device_pool[device] = {}
    # This first line just accounts for reserving space for the non data rows of this device.
    # The next lines account for space taken by previous devices.
    device_pool[device]["start"] = NUM_NON_DATA_ROWS + relative_row_data_start + \
        device_n * (NUM_NON_DATA_ROWS + relative_row_data_start + num_data_rows +
                    NUM_PADDING_CELLS)
    device_pool[device]["end"] = device_pool[device]["start"] + num_data_rows
    device_n += 1
  # Make note of where the next free row is, so that we can write more tables and graphs there.
  current_free_row = find_dict_val(device_pool, max) + 1 + NUM_PADDING_CELLS

  print(pool, device_pool)

  print(Fore.GREEN + "Writing CSVs to workbook...")

  wb = openpyxl.Workbook()

  ws_all = wb.active
  ws_all.title = "All Charge Integrals"

  # Write the all integrals headers.
  for device in csv_dict_ordered:
    # Note that these iterate over the pool, rather than the CSV dict. The effect of this is that
    # each device will have every row and column, regardless of whether it has data for it.
    for voltage in pool["row_mappings"]:
      for lifetime in pool["row_mappings"][voltage]:
        # - 1 because of not being 0-indexed.
        ws_all.cell(row=device_pool[device]["start"] + pool["row_mappings"][voltage][lifetime] - 1,
                    column=ALLINT_VOLTAGE_COL).value = voltage
        # Ditto.
        ws_all.cell(row=device_pool[device]["start"] + pool["row_mappings"][voltage][lifetime] - 1,
                    column=ALLINT_LIFETIME_COL).value = lifetime
    for density in pool["col_mappings"]:
      ws_all.cell(
          row=device_pool[device]["start"] - 1,
          column=NUM_NON_DATA_COLUMNS + pool["col_mappings"][density]).value \
          = "Charge ({})".format(density)

  # Populate the workbook with the data.
  print(Fore.CYAN + "CSVs:")
  # Define an equation for finding the difference between the charge at a point of the pulse,
  # and the charge at the pedestal of the pulse. This is basically the "effective" height.
  trans_difference = openpyxl.formula.translate.Translator("=B2-$B$2", origin="B2")
  # Define an equation for finding the integral at the point of a pulse, from the last point.
  trans_integral = openpyxl.formula.translate.Translator("=C3*(A3-A2)", origin="C3")
  for device in csv_dict_ordered:
    print(Fore.WHITE + "- {}".format(device))
    header_row = device_pool[device]["start"] - 1
    # Add a label for this table.
    # TODO: Format this!
    ws_all.cell(row=header_row - 1, column=1).value = "{}:".format(device)
    # Add a header for the voltages.
    # TODO: Format this!
    ws_all.cell(row=header_row, column=ALLINT_VOLTAGE_COL).value = "Voltage"
    # Add a header for the lifetimes.
    # TODO: Format this!
    ws_all.cell(row=header_row, column=ALLINT_LIFETIME_COL).value = "Lifetime"
    for voltage in csv_dict_ordered[device]:
      print(Fore.WHITE + "  - {}".format(voltage))
      for lifetime in csv_dict_ordered[device][voltage]:
        print(Fore.WHITE + "    - {}".format(lifetime))

        for density in csv_dict_ordered[device][voltage][lifetime]:
          # Write the data from the CSV file to a new sheet.
          path = csv_dict_ordered[device][voltage][lifetime][density]
          print(Fore.WHITE + "      - {}: ".format(density) +
                Style.BRIGHT + path)
          sheet_name = "{} {} {} {}".format(device, voltage, lifetime, density)
          ws = wb.create_sheet(sheet_name)
          with open(path) as f:
            reader = csv.reader(f)
            # Keep track of the row we're writing to so that we can translate the formulas.
            row_n = 0
            for row in reader:
              row_n += 1
              # Only add transient time and substrate current.
              ws.append([row[0], row[6]])
              if row_n > 1:
                ws["C{}".format(row_n)] = trans_difference.translate_formula(
                    "B{}".format(row_n))
                ws["D{}".format(row_n)] = trans_integral.translate_formula(
                    "C{}".format(row_n))
          # TODO: Is there a better name for this?
          ws["C1"] = "Difference from Start"
          ws["D1"] = "Integral"
          # The integral formula references the row above it, so it doesn't make sense to
          # have one in the first row.
          ws["D2"] = None
          ws["E1"] = "Sum of Integrals"
          ws["E2"] = "=SUM(D3:D{})".format(row_n)

          # Write the sum of integrals to the all integrals sheet.
          # - 1 because of not being 0-indexed.
          ws_all.cell(
              row=device_pool[device]["start"] +
              pool["row_mappings"][voltage][lifetime] - 1,
              column=NUM_NON_DATA_COLUMNS +
              pool["col_mappings"][density]).value = "='{}'!E2".format(sheet_name)

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Further processing data...")

  start_pc_row_data = device_pool["PC"]["start"]
  end_pc_row_data = device_pool["PC"]["end"]

  start_diode_row_data = device_pool["Diode"]["start"]

  start_nc_row_data = device_pool["PC_Norm"]["start"]
  header_row = start_nc_row_data - 1

  ws_all.cell(row=header_row - 1, column=1).value = "PC Normalized:"

  trans_normalize = openpyxl.formula.translate.Translator(
      "=C{}/C${}".format(start_pc_row_data, start_diode_row_data),
      origin="C{}".format(start_pc_row_data))

  ws_all.cell(row=current_free_row, column=1).value = "PC Graphs:"
  current_free_row += 1

  # Create Lifetime v. Charge graphs.
  # The - 2 is a very hacky solution to ignore the 100V and 1000V rows. In reality, this doesn't
  # seem to actually help the graph issues, but hey.
  x_values = openpyxl.chart.Reference(
      worksheet=ws_all,
      min_col=ALLINT_LIFETIME_COL,
      min_row=start_pc_row_data,
      max_row=end_pc_row_data - 2)
  left = True
  for density, col in pool["col_mappings"].items():
    column_is_empty = True
    for row in range(start_pc_row_data, end_pc_row_data + 1):
      if ws_all.cell(row=row, column=NUM_NON_DATA_COLUMNS + col).value is not None:
        column_is_empty = False
    if column_is_empty is True:
      continue

    values = openpyxl.chart.Reference(
        worksheet=ws_all,
        min_col=NUM_NON_DATA_COLUMNS + col,
        min_row=start_pc_row_data,
        max_row=end_pc_row_data - 2)

    chart = openpyxl.chart.LineChart()
    chart.title = "Charge vs. Lifetime ({})".format(density)
    # This style has a white main background, a light purple plot background, and a dark purple
    # line.
    chart.style = 38
    chart.width = 13
    chart.height = 7
    chart.anchor = ws_all.cell(row=current_free_row, column=1 if left else 6).coordinate
    chart.legend = None
    chart.x_axis.title = "Lifetime"
    chart.x_axis.scaling.logBase = 10
    chart.y_axis.title = "Charge"
    chart.add_data(values)
    chart.set_categories(x_values)
    ws_all.add_chart(chart)

    # Advance the "index" at which we are anchoring the chart.
    left = not left
    if left:
      # This is meant to be the height of the graph in cells.
      current_free_row += 13

    # Go back to working on the NC table.

    # This + 1 accounts for the range not including the last element.
    for row in range(relative_row_data_start, relative_row_data_end + 1):
      pc_cell = ws_all.cell(
          row=(start_pc_row_data - 1) + row,
          column=NUM_NON_DATA_COLUMNS + col)
      # print(col, row, pc_cell.value)
      if pc_cell.value is not None:
        ws_all.cell(
            row=(start_nc_row_data - 1) + row,
            column=NUM_NON_DATA_COLUMNS + col).value \
            = trans_normalize.translate_formula(pc_cell.coordinate)

  current_free_row += 13 + NUM_PADDING_CELLS
  ws_all.cell(row=current_free_row, column=1).value = "PC Normalized Graphs:"
  current_free_row += 1

  # Hardcoding data sources? Nooo, I would never!
  values = openpyxl.chart.Reference(
      worksheet=ws_all,
      min_col=6,
      min_row=23,
      max_row=27)

  chart = openpyxl.chart.LineChart()
  chart.title = "Charge vs. Lifetime ({})".format(density)
  # This style has a white main background, a light purple plot background, and a dark purple
  # line.
  chart.style = 38
  chart.width = 13
  chart.height = 7
  chart.anchor = ws_all.cell(row=current_free_row, column=1).coordinate
  chart.legend = None
  chart.x_axis.title = "Lifetime"
  chart.x_axis.scaling.logBase = 10
  chart.y_axis.title = "Charge"
  chart.add_data(values)
  chart.set_categories(x_values)
  ws_all.add_chart(chart)

  # We want to graph charge as a function of density, rather than lifetime/voltage. This is
  # markedly different than the charts we have made so far. In order to accomplish, here, we copy
  # the table to the new format, in a new sheet.

  ws_diode = wb.create_sheet(title="Diode Charge Integral Analysis", index=1)
  ws_diode.cell(row=1, column=DIODE_DENSITY_COL).value = "Density"
  start_col_data = 1
  # Keep track of the column we're writing to, so we know where to put the graph.
  current_col = start_col_data
  # Map what used to be voltage rows, to voltage columns. - 1 is necessary to get rid of a header
  # cell that is present for the rows but not the columns.
  for voltage in pool["row_mappings"]:
    current_col += 1
    for lifetime in pool["row_mappings"][voltage]:
      # We are only interested in this particular lifetime.
      if "1e-7" in lifetime:
        row_orig = pool["row_mappings"][voltage][lifetime]
        ws_diode.cell(row=1, column=current_col).value = "Charge ({})".format(voltage)
  end_col_data = current_col

  start_row_data = 1
  # Keep track of the row we're writing to, so we know what data to use for the graph.
  current_row = start_row_data
  # Note that this iterates in a different order than everywhere else, in doing density, *then*
  # voltage.
  for density in pool["col_mappings"]:
    current_row += 1
    col_orig = pool["col_mappings"][density]
    # For this, we don't want the "D=".
    row_contents = [density.split('=')[-1]]
    for voltage in pool["row_mappings"]:
      for lifetime in pool["row_mappings"][voltage]:
        # We are still only interested in this particular lifetime.
        if "1e-7" in lifetime:
          row_orig = pool["row_mappings"][voltage][lifetime]
          row_contents.append(
              ws_all.cell(
                  row=row_orig +
                  NUM_NON_DATA_ROWS,
                  column=col_orig +
                  NUM_NON_DATA_COLUMNS).value)
    ws_diode.append(row_contents)
  end_row_data = current_row

  # Now, we can create the chart.

  x_values = openpyxl.chart.Reference(
      worksheet=ws_diode,
      min_col=DIODE_DENSITY_COL,
      # For setting the category, we don't want to include the header.
      min_row=start_row_data + 1,
      max_row=end_row_data)

  values = openpyxl.chart.Reference(
      worksheet=ws_diode,
      min_col=DIODE_DENSITY_COL + 1,
      min_row=start_row_data,
      max_col=end_col_data,
      max_row=end_row_data)

  chart = openpyxl.chart.LineChart()
  chart.title = "Charge vs. Density"
  chart.style = 38
  chart.width = 25
  chart.height = 18
  # Add 1 to make a gap between the table and chart.
  chart.anchor = ws_diode.cell(row=start_row_data, column=end_col_data + 2).coordinate
  chart.x_axis.title = "Density"
  # It would be nice to fix up the scale so less space is wasted, but this doesn't seem to work
  # very well/at all.
  # chart.x_axis.scaling.min = ws_diode.cell(row=2, column=DIODE_DENSITY_COL).value
  chart.x_axis.scaling.logBase = 10
  chart.y_axis.scaling.logBase = 10
  chart.y_axis.title = "Charge"
  chart.add_data(values, titles_from_data=True)
  chart.set_categories(x_values)
  ws_diode.add_chart(chart)

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Adjusting column sizes...")

  # See: https://stackoverflow.com/a/35790441
  for ws in wb.worksheets:
    dims = {}
    # print("Fixing {}.".format(ws.title))
    for row in ws.rows:
      for cell in row:
        if cell.value:
          if cell.value.startswith('='):
            cell_len = 12
          else:
            # Add 1 to make things a little more roomy. Without this, the non-compilation sheets are
            # pretty cramped.
            cell_len = len(str(cell.value)) + 1
          dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), cell_len))
    # print(dims)
    for col, value in dims.items():
      ws.column_dimensions[col].width = value

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Writing Excel file...")

  wb.save("Charge Data.xlsx")

  print(Fore.GREEN + "Success.")

  os.sys.exit(0)


def main():
  """Entrypoint for subcommands.
  """

  colorama_init(autoreset=True)

  parser = argparse.ArgumentParser()
  subparsers = parser.add_subparsers()

  subparser_ls = subparsers.add_parser(
      "ls", help="Lists experiment directories.")
  subparser_ls.add_argument(
      "dir",
      help="Directory to read experiment directories from.",
      type=str,
      nargs="?",
      default=DEFAULT_DATA_DIR)
  subparser_ls.set_defaults(func=ls)

  subparser_export = subparsers.add_parser(
      "export", help="Exports an experiment to Excel.")
  subparser_export.add_argument(
      "exp", help="Directory or name of the experiment to export.", type=str)
  subparser_export.set_defaults(func=export)

  # TODO: Fix this (catching errors like this causes other issues).
  args = parser.parse_args()
  # Ask for forgiveness from the Python gods (https://stackoverflow.com/a/610923).
  # try:
  args.func(args)
  # except AttributeError:
  #   parser.print_help()


if __name__ == "__main__":
  main()
