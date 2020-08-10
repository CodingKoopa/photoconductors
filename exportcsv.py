#!/usr/bin/env python3

# Requires Python >= 3.8 for assignment within condition: https://www.python.org/dev/peps/pep-0572/.
# Also requires Python >= 3.6 for preserving order of inseration for dictionaries.
# Also requires Python >= 3.5 for recursive support in glob.
# pip install -r requirements.txt

# Compatibility:
# - The online version of Microsoft Excel works without any issues.
# - Google Sheets has issues with decimal places. The default level of precision is not enough to
# support the math done. This may be fixable.
# - LibreOffice Calc has issues with sheet references. It uses a different format than Excel and
# Google Sheets, but seems to convert the formula to all lowercase, breaking the name of the sheet
# being reference. It is not supported by this script.

import argparse
import csv
from glob import glob
import os
import re

from colorama import Fore, Style, init as colorama_init
import openpyxl

DEFAULT_DATA_DIR = "../data"
VOLTAGE_FILENAME_STR = "voltage"
LIFETIME_FILENAME_STR = "lifetime"
DENSITY_FILENAME_STR = "ehpdensity"
NUM_NON_DATA_COLUMNS = 2


def sort_alphanumeric(data):
  """See: https://stackoverflow.com/a/48030307
  """

  def convert(text): return int(text) if text.isdigit() else text.lower()
  def alphanum_key(key): return [convert(c)
                                 for c in re.split('([0-9]+)', key)]
  return sorted(data, key=alphanum_key)


def ls(args):
  """TODO
  """

  print(Fore.GREEN + "Listing experiments...")
  # Fetch top-level directories in the data directory.
  child_dirs = glob(args.dir + "/*/")
  # Only consider directories with CSV files in them.
  experiment_dirs = sort_alphanumeric(
      [d for d in child_dirs if glob(d + "/**/*.csv", recursive=True)])
  if experiment_dirs:
    print(Fore.CYAN + "Experiments:")
    for d in experiment_dirs:
      # The path must be normalized in order for basename to not return an empty string.
      print("- " + Fore.WHITE + os.path.basename(os.path.normpath(d)))
    print(Fore.GREEN + "Success.")
    os.sys.exit(0)
  else:
    print(Fore.RED + "No experiments found.")
    os.sys.exit(1)


def export(args):
  """TODO
  """

  print(Fore.GREEN + "Mapping CSV files to dictionary...")

  exp = args.exp
  # Provide support for only passing the name of the experiment.
  if not os.path.exists(exp):
    exp = os.path.join(DEFAULT_DATA_DIR, exp)
    if not os.path.exists(exp):
      print(Fore.RED + "Experiment \"{}\" not found.".format(args.exp))
      os.sys.exit(1)
  csv_list = glob(exp + "/**/*.csv", recursive=True)
  if not csv_list:
    print(Fore.RED + "No CSV files found.")
    os.sys.exit(1)

  print(Fore.CYAN + "CSV Files:")
  csv_dict = {}
  for f in csv_list:
    print(Fore.WHITE + Style.BRIGHT + f)
    # File names are expected to come in formats like
    # "PC_lifetime0.001_ehpdensity0.0001_time.csv". Here, we parse the information given,
    # delimited by the underscores.
    basename_parts = os.path.basename(f).split('_')

    # See: https://stackoverflow.com/a/30197797.
    device_str = next((s for s in basename_parts if "PC" in s), None)
    if not device_str:
      device_str = next(
          (s for s in basename_parts if "Diode" in s), None)
      if not device_str:
        print(
            Fore.RED +
            "Device name could not be extracted from file \"{}\". Looked for \"PC\" or \"Diode\" \
in file name. Skipping this file.".format(f))
        continue

    # This works in the same way as its voltage analog.
    voltage_str = next(
        (s for s in basename_parts if VOLTAGE_FILENAME_STR in s), None)
    if not voltage_str:
      print(
          Fore.RED +
          "Density could not be extracted from file \"{}\". Looked for \"{}\" in file name. \
Skipping this file." .format(f, VOLTAGE_FILENAME_STR))
      continue
    # TODO: what formatting to use?
    voltage_key = "{}V".format(
        int(voltage_str.replace(VOLTAGE_FILENAME_STR, '')))

    # Look for LIFETIME_FILENAME_STR in the file name to find the field, and then remove that
    # identifier string so that we are just left with the value of the field. For example,
    # "lifetime0.001" is transformed into "τ=1.00e-3".
    lifetime_str = next(
        (s for s in basename_parts if LIFETIME_FILENAME_STR in s), None)
    if not lifetime_str:
      print(
          Fore.RED +
          "Lifetime could not be extracted from file \"{}\". Looked for \"{}\" in file name. \
Skipping this file." .format(f, LIFETIME_FILENAME_STR))
      continue
    # Force scientific notation so that the dictionary key can be properly sorted as a float.
    lifetime_key = "τ={:0.2e}".format(
        float(lifetime_str.replace(LIFETIME_FILENAME_STR, '')))

    # This works in the same way as its lifetime analog.
    density_str = next(
        (s for s in basename_parts if DENSITY_FILENAME_STR in s), None)
    if not density_str:
      print(
          Fore.RED +
          "Density could not be extracted from file \"{}\". Looked for \"{}\" in file name. \
Skipping this file." .format(f, DENSITY_FILENAME_STR))
      continue
    density_key = "D={:0.2e}".format(
        float(density_str.replace(DENSITY_FILENAME_STR, '')))

    # device_str is already usable as a key.
    if device_str not in csv_dict:
      csv_dict[device_str] = {}
    # Initialize lifetime and voltage dictionaries as needed.
    if voltage_key not in csv_dict[device_str]:
      csv_dict[device_str][voltage_key] = {}
    if lifetime_key not in csv_dict[device_str][voltage_key]:
      csv_dict[device_str][voltage_key][lifetime_key] = {}
    csv_dict[device_str][voltage_key][lifetime_key][density_key] = f

  if not csv_dict:
    print(Fore.RED + "Unable to create a CSV dictionary.")
    os.sys.exit(1)
  print(Fore.GREEN + "Success.")

  # This is necessary because alphanumeric sorting of the file names may not account for
  # scientific notation.
  print(Fore.GREEN + "Ordering dictionary...")

  # This (https://stackoverflow.com/a/47882384) SO answer explains how to recursively sort a
  # dictionary. I have modified this to sort the root entries as normal, but sort further nested
  # entries as floats/ints. This also accounts for the equal sign in the key. This was annoying to
  # figure out.
  def sort_dict(dictionary, level=0):
    # print("sort_dict({})".format(root))
    result = {}
    if level == 0:
      key = None
    elif level == 1:
      key = (lambda x: int(x[0].split('V')[0]))
    else:
      key = (lambda x: float(x[0].split('=')[-1]))
    for k, v in sorted(dictionary.items(), key=key):
      # print("{}:{}".format(k, v))
      if isinstance(v, dict):
        result[k] = sort_dict(v, level + 1)
      else:
        result[k] = v
    return result

  csv_dict_ordered = sort_dict(csv_dict)

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Writing CSVs to workbook...")

  wb = openpyxl.Workbook()

  # TODO describe this
  ws_comp = wb.active
  ws_comp.title = "Compilation"

  # This is a dictionary used to keep track of "tables" on the compilation sheet. It will be filled
  # with a nested dictionary for each device, which contains mappings of properties to rows/columns.
  device_tables = {}
  # Note that these mean "current" as in "right now", not in the electrical sense.
  current_header_row = 2
  current_free_row = 3
  # If changing this, be sure to change the value at the end of the for loop!
  current_free_col = 1 + NUM_NON_DATA_COLUMNS
  last_pc_col = 1

  # Populate the workbook with the data.
  print(Fore.CYAN + "CSVs:")
  # Define an equation for finding the difference between the current at a point of the pulse,
  # and the current at the pedestal of the pulse. This is basically the "effective" height.
  trans_difference = openpyxl.formula.translate.Translator("=B2-$B$2", origin="B2")
  # Define an equation for finding the integral at the point of a pulse, from the last point.
  trans_integral = openpyxl.formula.translate.Translator("=C3*(A3-A2)", origin="C3")
  for device in csv_dict_ordered:
    print(Fore.WHITE + "- {}".format(device))
    # Add a device table.
    device_tables[device] = {"voltage_lifetime_to_row": {},
                             "density_to_col": {}}
    # Add a label for this table.
    ws_comp.cell(row=current_header_row - 1, column=1).value = "{}:".format(device)
    # Add a header for the voltages.
    # TODO: Format this!
    ws_comp.cell(row=current_header_row, column=1).value = "Voltage"
    # Add a header for the lifetimes.
    # TODO: Format this!
    ws_comp.cell(row=current_header_row, column=2).value = "Lifetime"
    for voltage in csv_dict_ordered[device]:
      print(Fore.WHITE + "  - {}".format(voltage))
      if voltage not in device_tables[device]["voltage_lifetime_to_row"]:
        device_tables[device]["voltage_lifetime_to_row"][voltage] = {}
      for lifetime in csv_dict_ordered[device][voltage]:
        print(Fore.WHITE + "    - {}".format(lifetime))

        # Check if there is a row allocated for this voltage lifetime, in the device table.
        if lifetime not in device_tables[device]["voltage_lifetime_to_row"][voltage]:
          device_tables[device]["voltage_lifetime_to_row"][voltage][lifetime] = current_free_row
          ws_comp.cell(row=device_tables[device]["voltage_lifetime_to_row"]
                       [voltage][lifetime], column=1).value = voltage
          ws_comp.cell(row=device_tables[device]["voltage_lifetime_to_row"]
                       [voltage][lifetime], column=2).value = lifetime.split('=')[-1]
          # print("Allocating row {} for lifetime {}.".format(current_free_row, lifetime))
          current_free_row += 1

        for density in csv_dict_ordered[device][voltage][lifetime]:
          # Write the data from the CSV file to a new sheet.
          path = csv_dict_ordered[device][voltage][lifetime][density]
          print(Fore.WHITE + "      - {}: ".format(density) +
                Style.BRIGHT + path)
          sheet_name = "{} {} {} {}".format(device, voltage, lifetime, density)
          ws = wb.create_sheet(sheet_name)
          with open(path) as f:
            reader = csv.reader(f)
            row_n = 0
            for row in reader:
              row_n += 1
              # Only add transient time and substrate current.
              ws.append([row[0], row[6]])
              if row_n > 1:
                ws["C{}".format(row_n)] = trans_difference.translate_formula(
                    "B{}".format(row_n))
                ws["D{}".format(row_n)] = trans_integral.translate_formula(
                    "C{}".format(row_n))
          ws["C1"] = "Difference from Start"
          ws["D1"] = "Integral"
          # The integral formula references the row above it, so it doesn't make sense to
          # have one in the first row.
          ws["D2"] = None
          ws["E1"] = "Sum of Integrals"
          ws["E2"] = "=SUM(D3:D{})".format(row_n)

          # See the lifetime analog of this.
          if density not in device_tables[device]["density_to_col"]:
            device_tables[device]["density_to_col"][density] = current_free_col
            ws_comp.cell(
                row=current_header_row,
                column=device_tables[device]["density_to_col"][density]).value = \
                "Current (D={})".format(density.split('=')[-1])
            # print("Allocating column {} for density {}.".format(current_free_col, density))
            current_free_col += 1

          # Write the sum of integrals to the compilation.
          # NB: LibreOffice uses "." for sheet references. It still doesn't seem to work, though, as
          # the equation seems to get converted to lowercase.
          ws_comp.cell(
              row=device_tables[device]["voltage_lifetime_to_row"][voltage][lifetime],
              column=device_tables[device]["density_to_col"][density]).value = \
              "='{}'!E2".format(sheet_name)
    if device == "PC":
      last_pc_col = max(last_pc_col, current_free_col - 1)
    # Columns may be reused from table to table, for different purposes.
    current_free_col = 1 + NUM_NON_DATA_COLUMNS
    # Make the header row start 2 below to make room for an empty space, and "$DEVICE:"
    current_header_row = current_free_row + 2
    current_free_row = current_header_row + 1

  # Move the current free row back to account for the manipulation done at the end of the loop, and
  # that we don't need to account for a header row.
  current_free_row = current_header_row - 1
  # Make this variable invalid to use (as we are done writing tables with headers).
  current_header_row = 0

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Writing charts to workbook...")

  def find_dict_val(dic, f):
    first_sub_dic = dic[next(iter(dic))]
    # Initialize to first value in dict.
    val = first_sub_dic[f(first_sub_dic, key=first_sub_dic.get)]
    for sub_dic_key in dic:
      sub_dic = dic[sub_dic_key]
      val = f(val, sub_dic[f(sub_dic, key=sub_dic.get)])
    return val

  if "PC" in device_tables:
    device_table = device_tables["PC"]
    # See: https://stackoverflow.com/a/280156.
    start_row_data = find_dict_val(device_table["voltage_lifetime_to_row"], min)
    end_row_data = find_dict_val(device_table["voltage_lifetime_to_row"], max)

    # This is redundant with the Diode charting code. Not sure how/if I want to fix that though, as
    # the code for different charts/tables are meant to be fairly independent.
    if "Diode" in device_tables:
      start_row_data_diode = find_dict_val(device_tables["Diode"]["voltage_lifetime_to_row"], min)
    else:
      print(Fore.RED + "Unable to find Diode table.")
      os.sys.exit(1)

    # Create Normalized Charges table. This gets tacked onto the PC data table, so we will be
    # reusing those row numbers, but shifting columns as appropriate. No + 1 is necessary as we will
    # offset as necessary later.
    start_col_nc = last_pc_col

    trans_normalize = openpyxl.formula.translate.Translator(
        "=C{}/C${}".format(start_row_data, start_row_data_diode),
        origin="C{}".format(start_row_data))
    # (This will be returned to in a little bit.)

    # Create Lifetime v. Current graphs.
    x_values = openpyxl.chart.Reference(
        worksheet=ws_comp,
        min_col=2,
        min_row=start_row_data,
        max_row=end_row_data)
    left = True
    for density, col in device_table["density_to_col"].items():
      values = openpyxl.chart.Reference(
          worksheet=ws_comp,
          min_col=col,
          min_row=start_row_data,
          max_row=end_row_data)

      chart = openpyxl.chart.LineChart()
      chart.title = "Lifetime vs. Current ({})".format(density)
      # This style has a white main background, a light purple plot background, and a dark purple
      # line.
      chart.style = 38
      chart.width = 13
      chart.height = 7
      chart.anchor = ws_comp.cell(row=current_free_row, column=1 if left else 6).coordinate
      chart.legend = None
      chart.x_axis.title = "Lifetime"
      chart.x_axis.scaling.logBase = 10
      chart.y_axis.title = "Current"
      chart.add_data(values)
      chart.set_categories(x_values)
      ws_comp.add_chart(chart)

      # Advance the "index" at which we are anchoring the chart.
      left = not left
      if left:
        # This is meant to be the height of the graph in cells.
        current_free_row += 13

      # Go back to working on the NC table.
      ws_comp.cell(
          row=start_row_data - 1,
          column=start_col_nc + (col - NUM_NON_DATA_COLUMNS)).value = "Norm. {}".format(
          ws_comp.cell(
              row=start_row_data - 1,
              column=col).value)
      # This + 1 accounts for the range not including the last element.
      for row in range(start_row_data, end_row_data + 1):
        # - 2 to account for the two columns of voltage/lifetime.
        ws_comp.cell(row=row, column=start_col_nc + (col - NUM_NON_DATA_COLUMNS)).value = \
            trans_normalize.translate_formula(ws_comp.cell(row=row, column=col).coordinate)
  else:
    print(Fore.RED + "Unable to find PC table.")
    os.sys.exit(1)

  # TODO: Create "" graph
  if "Diode" in device_tables:
    device_table = device_tables["Diode"]

    # We want to graph current as a function of density, rather than lifetime/voltage. This is
    # markedly different than the charts we have made so far. In order to accomplish, here, we copy
    # the table to the new format, in a new worksheet.

    ws_diode = wb.create_sheet(title="TODO", index=0)
    ws_diode.cell(row=1, column=1).value = "Density"
    # Map what used to be voltage rows, to voltage columns. - 1 is necessary to get rid of a header
    # cell that is present for the rows but not the cells.
    for voltage in device_table["voltage_lifetime_to_row"]:
      # Take the first element in the dictionary because we don't care about lifetimes, and there
      # should be only one.
      row_orig = device_table["voltage_lifetime_to_row"][voltage][next(
          iter(device_table["voltage_lifetime_to_row"][voltage]))]
      ws_diode.cell(row=1, column=row_orig - 1).value = "Current ({})".format(voltage)

    # Note that this iterates in a different order than everywhere else, in doing density, *then*
    # voltage.
    for density in device_table["density_to_col"]:
      col_orig = device_table["density_to_col"][density]
      row_contents = [density.split('=')[-1]]
      for voltage in device_table["voltage_lifetime_to_row"]:
        # Same as above.
        row_orig = device_table["voltage_lifetime_to_row"][voltage][next(
            iter(device_table["voltage_lifetime_to_row"][voltage]))]
        row_contents.append(ws_comp.cell(row=row_orig, column=col_orig).value)
      ws_diode.append(row_contents)

    # Now, we can create the new charts.
  else:
    print(Fore.RED + "Unable to find PC table.")
    os.sys.exit(1)

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Adjusting column sizes...")

  # See: https://stackoverflow.com/a/35790441
  for ws in wb.worksheets:
    dims = {}
    # print("Fixing {}.".format(ws.title))
    for row in ws.rows:
      for cell in row:
        if cell.value:
          if cell.value.startswith('='):
            cell_len = 12
          else:
            # Add 1 to make things a little more roomy. Without this, the non-compilation sheets are
            # pretty cramped.
            cell_len = len(str(cell.value)) + 1
          dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), cell_len))
    # print(dims)
    for col, value in dims.items():
      ws.column_dimensions[col].width = value

  print(Fore.GREEN + "Success.")

  print(Fore.GREEN + "Writing Excel file...")

  wb.save("Current Data.xlsx")

  print(Fore.GREEN + "Success.")

  os.sys.exit(0)


def main():
  """Entrypoint for subcommands.
  """

  colorama_init(autoreset=True)

  parser = argparse.ArgumentParser()
  subparsers = parser.add_subparsers()

  subparser_ls = subparsers.add_parser(
      "ls", help="Lists experiment directories.")
  subparser_ls.add_argument(
      "dir",
      help="Directory to read experiment directories from.",
      type=str,
      nargs="?",
      default=DEFAULT_DATA_DIR)
  subparser_ls.set_defaults(func=ls)

  subparser_export = subparsers.add_parser(
      "export", help="Exports an experiment to Excel.")
  subparser_export.add_argument(
      "exp", help="Directory or name of the experiment to export.", type=str)
  subparser_export.set_defaults(func=export)

  args = parser.parse_args()
  # Ask for forgiveness from the Python gods (https://stackoverflow.com/a/610923).
  # try:
  args.func(args)
  # except AttributeError:
  #   parser.print_help()


if __name__ == "__main__":
  main()
